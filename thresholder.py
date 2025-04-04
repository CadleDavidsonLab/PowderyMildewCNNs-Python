__author__ = "Dani Martinez"
__copyright__ = "Copyright 2025, Moblanc Robotics & Cornell University"
__credits__ = ["Dani Martinez"]
__license__ = "Apache 2.0"
__version__ = "0.5"
__maintainer__ = "Dani Martinez"
__email__ = "dani.martinez@moblancrobotics.com"
__status__ = "Production"

from pathlib import Path
import os
import argparse

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import numpy as np
import msgpack
import msgpack_numpy as mnp
mnp.patch()


# Useful colored strings
hwarning = "\033[93m[WARNING]\033[0m: "
herror = "\033[91m[ERROR]\033[0m: "


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Blackbird Results Thresholder")
    parser.add_argument(
        "res_path",
        metavar="<RES_PATH>",
        help="Path to experiment results as a *.msgpack file.",
    )
    parser.add_argument(
        "-o",
        "--out",
        metavar="<OUT_XLSX>",
        required=False,
        default="results.xlsx",
        help="Path where the Excel file will be placed and its filename (incl. .xlsx extension).",
    )
    parser.add_argument(
        "-lo",
        "--low",
        help="Specify lower threshold between 0 and 1",
        metavar="<LOW_TH>",
        default=0.5,
        type=float,
        required=False,
    )
    parser.add_argument(
        "-hi",
        "--high",
        help="Specify higher threshold between 0 and 1",
        metavar="<HIGH_TH>",
        default=0.5,
        type=float,
        required=False,
    )
    args = parser.parse_args()

    if not os.path.exists(args.res_path):
        print(herror + " Specified results path '" + args.res_path + "' does not exist!")
        exit()

    if args.low > args.high:
        print(herror + " Specified tresholds are not valid!")
        exit()

    res_path = Path(args.res_path)

    # Read msgpack file
    with open(res_path, "rb") as data_file:
        byte_data = data_file.read()
        results = msgpack.unpackb(byte_data)


    print(f" ** Thresholding results... Low_th: {args.low}, High_th: {args.high}")

    # Create Excel sheet
    wb = openpyxl.Workbook()

    # Get unique tray IDs
    trays_sheets = {}
    for d in results:
        for t in results[d]:
            if t not in trays_sheets:
                trays_sheets[t] = wb.create_sheet(t)

    dates_ids = results.keys()

    # Delete default sheet
    del wb['Sheet']

    #print(trays_sheets)
    thin_border = Side(border_style="thin", color="000000")
    thick_border = Side(border_style="thick", color="000000")
    header_color = PatternFill(fgColor="CCCCCC", fill_type = "solid")

    inf_color = PatternFill(fgColor="F5B7B1", fill_type = "solid")
    clr_color = PatternFill(fgColor="B0E0E6", fill_type = "solid")

    tick_left = Border(top=thin_border, left=thick_border, right=thin_border, bottom=thin_border)
    tick_right = Border(top=thin_border, left=thin_border, right=thick_border, bottom=thin_border)
    n_processed = 0

    # Initialize row and column headers (dates & sampleIDs)
    for t in trays_sheets.keys():

        # Apply thin borders in all sheet
        for row in trays_sheets[t]["A1:"+get_column_letter((len(dates_ids)*4)+1)+"353"]:
            for cell in row:
                cell.border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)

        # Initialize sample ids with N/A
        trays_sheets[t]["A1"] = "SampleID"
        trays_sheets[t].column_dimensions["A"].width = 30
        trays_sheets[t].merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        trays_sheets[t]["A1"].alignment = Alignment(horizontal='center',vertical="center")
        trays_sheets[t]["A1"].font = Font(bold=True)
        trays_sheets[t]["A1"].fill = header_color
        

        # Intialize all sampleIDs as N/A
        for r in range(3,354):
            trays_sheets[t].cell(row=r, column=1).value = "N/A"

            # Apply thick borders separating dates
            for j, d in enumerate(dates_ids):
                trays_sheets[t].cell(row=r, column=(j*4)+2).border = tick_left
                trays_sheets[t].cell(row=r, column=(j*4)+5).border = tick_right

        # Initialize date headers (all trays same even if some timepoints don't have a specific tray)
        for j, d in enumerate(dates_ids):
            trays_sheets[t].cell(row=1, column=(j*4)+2).value = d
            trays_sheets[t].cell(row=1, column=(j*4)+2).font = Font(bold=True)
            trays_sheets[t].cell(row=1, column=(j*4)+2).fill = header_color
            trays_sheets[t].merge_cells(start_row=1, start_column=(j*4)+2, end_row=1, end_column=(j*4)+5)
            trays_sheets[t].cell(row=1, column=(j*4)+2).border =  tick_left
            trays_sheets[t].cell(row=1, column=(j*4)+5).border =  tick_right

            trays_sheets[t].column_dimensions[get_column_letter((j*4)+2)].width = 5
            trays_sheets[t].column_dimensions[get_column_letter((j*4)+3)].width = 5
            trays_sheets[t].column_dimensions[get_column_letter((j*4)+4)].width = 5
            trays_sheets[t].column_dimensions[get_column_letter((j*4)+5)].width = 5
            trays_sheets[t][get_column_letter((j*4)+2)+"1"].alignment = Alignment(horizontal='center')

            trays_sheets[t].cell(row=2, column=(j*4)+2).value = "%"
            trays_sheets[t].cell(row=2, column=(j*4)+2).alignment = Alignment(horizontal='center')
            trays_sheets[t].cell(row=2, column=(j*4)+2).fill = inf_color
            trays_sheets[t].cell(row=2, column=(j*4)+2).font = Font(bold=True)
            trays_sheets[t].cell(row=2, column=(j*4)+2).border = tick_left

            trays_sheets[t].cell(row=2, column=(j*4)+3).value = "INF"
            trays_sheets[t].cell(row=2, column=(j*4)+3).alignment = Alignment(horizontal='center')
            trays_sheets[t].cell(row=2, column=(j*4)+3).fill = inf_color
            trays_sheets[t].cell(row=2, column=(j*4)+3).font = Font(bold=True)

            trays_sheets[t].cell(row=2, column=(j*4)+4).value = "CLR"
            trays_sheets[t].cell(row=2, column=(j*4)+4).alignment = Alignment(horizontal='center')
            trays_sheets[t].cell(row=2, column=(j*4)+4).fill = clr_color
            trays_sheets[t].cell(row=2, column=(j*4)+4).font = Font(bold=True)
            
            trays_sheets[t].cell(row=2, column=(j*4)+5).value = "ALL"
            trays_sheets[t].cell(row=2, column=(j*4)+5).alignment = Alignment(horizontal='center')
            trays_sheets[t].cell(row=2, column=(j*4)+5).fill = header_color
            trays_sheets[t].cell(row=2, column=(j*4)+5).font = Font(bold=True)
            trays_sheets[t].cell(row=2, column=(j*4)+5).border = tick_right

            if t in results[d]:
                for i, s in enumerate(results[d][t]):
                    if s is not None:

                        # NOTE: This cell assignment will be overwritten multiple times but it's OK
                        trays_sheets[t].cell(row=i+3, column=1).value = s[0]

                        # Compute sample results
                        if s[1] is not None:
                            n_analyzed_subim = np.count_nonzero(~np.isnan(s[1]))
                            n_infected = np.count_nonzero(s[1] >= args.high)
                            n_clear = np.count_nonzero(s[1] < args.low)
                            inf_percent = int(n_infected / (n_infected+n_clear) * 100)
                        else: # No sample detected
                            n_analyzed_subim = 0
                            n_infected = 0
                            n_clear = 0
                            inf_percent = 0
                            trays_sheets[t].cell(row=i+3, column=(j*4)+2).fill = inf_color
                            trays_sheets[t].cell(row=i+3, column=(j*4)+3).fill = inf_color
                            trays_sheets[t].cell(row=i+3, column=(j*4)+4).fill = inf_color
                            trays_sheets[t].cell(row=i+3, column=(j*4)+5).fill = inf_color

                        trays_sheets[t].cell(row=i+3, column=(j*4)+2).value = inf_percent
                        trays_sheets[t].cell(row=i+3, column=(j*4)+2).alignment = Alignment(horizontal='center')
                        trays_sheets[t].cell(row=i+3, column=(j*4)+3).value = n_infected
                        trays_sheets[t].cell(row=i+3, column=(j*4)+3).alignment = Alignment(horizontal='center')
                        trays_sheets[t].cell(row=i+3, column=(j*4)+4).value = n_clear
                        trays_sheets[t].cell(row=i+3, column=(j*4)+4).alignment = Alignment(horizontal='center')
                        trays_sheets[t].cell(row=i+3, column=(j*4)+5).value = n_analyzed_subim
                        trays_sheets[t].cell(row=i+3, column=(j*4)+5).alignment = Alignment(horizontal='center')

                        n_processed += 1
                        print(f" ** {n_processed} samples processed.",end='\r')

        trays_sheets[t].freeze_panes = trays_sheets[t].cell(row=3, column=(len(dates_ids)*4)+2)

    print("\n ** Done")

    wb.save(args.out)
    print("Resulting Excel file saved to:", os.path.abspath(args.out))
